import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from openpyxl import load_workbook

# --- CONFIGURATION ---
url = 'https://www.shyp.gov.cn/shypq/yqyw-wb-gtjzl-gsgg-fags/index.html'
txt_filename = 'crawled_titles.txt'
excel_filename = '模版.xlsx'
sheet_name_to_update = 'shanghai'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

# --- PART 1: DATA COLLECTION AND PROCESSING (DO NOT CHANGE) ---
try:
    print(f"Connecting to {url}...")
    # ... (The entire code for crawling and extracting data remains the same as before)
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    response.encoding = 'utf-8'

    soup = BeautifulSoup(response.text, 'html.parser')
    list_items = soup.select('ul.uli16.nowrapli.padding-top-10.list-date.border li')
    
    extracted_data = []
    raw_titles = []
    common_suffix = "既有多层住宅加装电梯工程规划设计方案公示"

    print("Starting data collection and extraction...")
    for item in list_items:
        date_tag = item.find('span', class_='time')
        title_tag = item.find('a')
        if date_tag and title_tag:
            full_title = title_tag.get_text(strip=True)
            raw_titles.append(full_title)
            
            district, address = "Unknown", "Unknown"
            if '区' in full_title:
                district_end_index = full_title.find('区') + 1
                district = full_title[:district_end_index]
                address_part = full_title[district_end_index:]
                address = address_part.replace(common_suffix, "").strip()

            extracted_data.append({
                '区': district, '地址': address, '公告日期': date_tag.get_text(strip=True)
            })

    if raw_titles:
        with open(txt_filename, 'w', encoding='utf-8') as f:
            f.write('\n'.join(raw_titles))
        print(f"Complete! Saved {len(raw_titles)} raw titles to file '{txt_filename}'")


    # --- PART 2: WRITE ACCURATELY TO SPECIFIC EXCEL CELLS ---
    if extracted_data:
        new_df = pd.DataFrame(extracted_data)[['区', '地址', '公告日期']]

        # Load existing workbook
        book = load_workbook(excel_filename)
        
        # Get the sheet to update
        if sheet_name_to_update in book.sheetnames:
            sheet = book[sheet_name_to_update]
        else:
            sheet = book.create_sheet(sheet_name_to_update)
            # If it's a new sheet, write headers starting from column B
            headers_list = list(new_df.columns)
            sheet.cell(row=1, column=2, value=headers_list[0]) # Column B
            sheet.cell(row=1, column=3, value=headers_list[1]) # Column C
            sheet.cell(row=1, column=4, value=headers_list[2]) # Column D

        # Find the next empty row to start writing data
        start_row = sheet.max_row + 1

        # Loop through each new data row and write to specific cells
        print(f"Writing new data to sheet '{sheet_name_to_update}' starting from column B...")
        for i, row_data in new_df.iterrows():
            current_row = start_row + i
            # Write data to columns B, C, D
            sheet.cell(row=current_row, column=2, value=row_data['区'])      # Column B (index 2)
            sheet.cell(row=current_row, column=3, value=row_data['地址'])    # Column C (index 3)
            sheet.cell(row=current_row, column=4, value=row_data['公告日期']) # Column D (index 4)
        
        # Save the entire workbook
        book.save(excel_filename)
        
        print("Complete! File has been updated successfully.")
    else:
        print("No extracted data found to save to Excel file.")

except Exception as e:
    print(f"An error occurred: {e}")